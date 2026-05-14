from __future__ import annotations

from datetime import date

import pandas as pd
from openpyxl import load_workbook

from boardbudget.excel_store import create_new_board_file, load_board, save_board, write_generated_sheets
from boardbudget.models import Activity, Assignment, BoardData, BoardSettings, Person


def test_create_and_load_board_file(tmp_path) -> None:
    path = tmp_path / "board.xlsx"

    create_new_board_file(path, "My Board", date(2026, 5, 13))
    loaded = load_board(path)

    assert loaded.settings.board_name == "My Board"
    assert loaded.settings.start_date == date(2026, 5, 13)
    assert [p.person_id for p in loaded.people] == ["P1", "P2"]


def test_save_board_and_write_generated_sheets(tmp_path) -> None:
    path = tmp_path / "board.xlsx"
    board = BoardData(
        settings=BoardSettings("Excel", date(2026, 5, 13)),
        people=[Person("P1", "One", 8, True)],
        activities=[Activity("A1", 1, "Work", 8, 8, "PLANNED")],
        assignments=[Assignment("A1", "P1")],
    )

    save_board(path, board)
    write_generated_sheets(
        path,
        pd.DataFrame([{"date": "2026-05-13", "person_id": "P1", "person_name": "One", "activity_id": "A1", "activity_name": "Work", "hours": 8}]),
        pd.DataFrame([{"metric": "total_allocated_hours", "value": 8}]),
        pd.DataFrame([{"level": "INFO", "code": "OK", "message": "ok"}]),
        pd.DataFrame([{"activity_id": "A1", "estimated_value": 400}]),
    )

    workbook = load_workbook(path, read_only=True)
    assert {"01_Board", "02_People", "03_Activities", "04_Assignments", "05_Calendar", "06_Dashboard", "07_Warnings", "08_Activity_Economics"}.issubset(
        set(workbook.sheetnames)
    )
    loaded = load_board(path)
    assert loaded.activities[0].activity_id == "A1"


def test_load_old_activity_sheet_defaults_v2_columns(tmp_path) -> None:
    path = tmp_path / "old.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"key": "board_name", "value": "Old"},
                {"key": "start_date", "value": "2026-05-13"},
                {"key": "hours_per_day", "value": 8},
                {"key": "working_days", "value": "MON,TUE,WED,THU,FRI"},
            ]
        ).to_excel(writer, sheet_name="01_Board", index=False)
        pd.DataFrame([{"person_id": "P1", "name": "One", "hours_per_day": 8, "active": True}]).to_excel(writer, sheet_name="02_People", index=False)
        pd.DataFrame(
            [{"activity_id": "A1", "order": 1, "name": "Work", "estimated_hours": 40, "max_hours_per_day": 8, "status": "PLANNED", "notes": ""}]
        ).to_excel(writer, sheet_name="03_Activities", index=False)
        pd.DataFrame([{"activity_id": "A1", "person_id": "P1"}]).to_excel(writer, sheet_name="04_Assignments", index=False)

    loaded = load_board(path)

    assert loaded.activities[0].estimated_days == 5
    assert loaded.activities[0].daily_price == 0
