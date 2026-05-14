from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from boardbudget.dashboard import build_calendar_dataframe, build_calendar_pivot_dataframe, build_dashboard_dataframe
from boardbudget.export import prepare_board_download_file
from boardbudget.models import Absence, Activity, Assignment, BoardData, BoardSettings, Person
from boardbudget.planner_engine import calculate_plan
from boardbudget.sample_data import create_sample_board


def board(absences: list[Absence] | None = None, people: list[Person] | None = None, activities: list[Activity] | None = None, assignments: list[Assignment] | None = None) -> BoardData:
    return BoardData(
        settings=BoardSettings("Absences", date(2026, 5, 13)),
        people=people or [Person("P1", "One", 8, True, 400)],
        activities=activities or [Activity("A1", 1, "Work", 8, 8, "PLANNED", "", daily_price=500)],
        assignments=assignments or [Assignment("A1", "P1")],
        absences=absences or [],
    )


def test_absence_x8_blocks_allocation() -> None:
    result = calculate_plan(board(absences=[Absence(date(2026, 5, 13), "P1", "X", 8, "ferie")]))

    assert [allocation.date for allocation in result.allocations] == [date(2026, 5, 14)]


def test_absence_x4_reduces_capacity() -> None:
    result = calculate_plan(board(absences=[Absence(date(2026, 5, 13), "P1", "X", 4, "mezza")]))

    assert [(allocation.date, allocation.hours) for allocation in result.allocations] == [(date(2026, 5, 13), 4), (date(2026, 5, 14), 4)]


def test_tentative_absence_is_ignored_by_planner() -> None:
    result = calculate_plan(board(absences=[Absence(date(2026, 5, 13), "P1", "?", 8, "forse")]))

    assert [(allocation.date, allocation.hours) for allocation in result.allocations] == [(date(2026, 5, 13), 8)]


def test_absence_does_not_affect_other_person() -> None:
    data = board(
        people=[Person("P1", "One", 8, True, 400), Person("P2", "Two", 8, True, 300)],
        activities=[Activity("A1", 1, "Shared", 8, 8, "PLANNED", "", daily_price=500)],
        assignments=[Assignment("A1", "P1"), Assignment("A1", "P2")],
        absences=[Absence(date(2026, 5, 13), "P1", "X", 8, "")],
    )
    result = calculate_plan(data)

    assert [(allocation.person_id, allocation.date, allocation.hours) for allocation in result.allocations] == [("P2", date(2026, 5, 13), 8)]


def test_half_day_absence_cost_is_based_on_allocated_hours() -> None:
    data = board(absences=[Absence(date(2026, 5, 13), "P1", "X", 4, "")])
    result = calculate_plan(data)
    metrics = dict(zip(build_dashboard_dataframe(result, data, today=date(2026, 5, 13))["metric"], build_dashboard_dataframe(result, data, today=date(2026, 5, 13))["value"]))

    assert metrics["delivered_cost_until_today"] == 200


def test_board_calendar_includes_absence_text() -> None:
    data = board(absences=[Absence(date(2026, 5, 13), "P1", "X", 8, "")])
    result = calculate_plan(data)
    calendar_df = build_calendar_dataframe(result, data)
    pivot = build_calendar_pivot_dataframe(calendar_df, data)

    assert "Absence 8h" in pivot[pivot["date"] == "2026-05-13"].iloc[0]["P1"]


def test_activity_economics_dashboard_includes_done_and_excludes_cancelled() -> None:
    data = board(
        activities=[
            Activity("A1", 1, "Planned", 8, 8, "PLANNED", "", daily_price=500),
            Activity("A2", 2, "Done", 16, 8, "DONE", "", daily_price=400),
            Activity("A3", 3, "Cancelled", 80, 8, "CANCELLED", "", daily_price=1000),
        ],
        assignments=[Assignment("A1", "P1")],
    )
    result = calculate_plan(data)
    metrics = dict(zip(build_dashboard_dataframe(result, data, today=date(2026, 5, 13))["metric"], build_dashboard_dataframe(result, data, today=date(2026, 5, 13))["value"]))

    assert metrics["total_estimated_value"] == 1300


def test_simplified_dashboard_metrics_exist() -> None:
    data = board()
    result = calculate_plan(data)
    metrics = set(build_dashboard_dataframe(result, data, today=date(2026, 5, 13))["metric"])

    assert {"total_estimated_value", "estimated_delivery_cost", "delivered_cost_until_today", "if_finished_today.theoretical_saving", "expected_margin"}.issubset(metrics)


def test_export_includes_absences_and_board_calendar(tmp_path) -> None:
    path = tmp_path / "sample.xlsx"
    create_sample_board(path)

    workbook = load_workbook(BytesIO(prepare_board_download_file(path)), read_only=True)

    assert {"04_Absences", "05_Calendar", "06_Dashboard", "07_Warnings", "08_Activity_Economics", "09_Board_Calendar", "10_Person_Economics"}.issubset(set(workbook.sheetnames))
    assert workbook["09_Board_Calendar"].max_row > 1

