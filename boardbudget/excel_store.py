from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .config import (
    ACTIVITIES_SHEET,
    ASSIGNMENTS_SHEET,
    ACTIVITY_ECONOMICS_SHEET,
    BOARD_SHEET,
    CALENDAR_SHEET,
    DASHBOARD_SHEET,
    DEFAULT_HOURS_PER_DAY,
    DEFAULT_WORKING_DAYS,
    PEOPLE_SHEET,
    PERSON_ECONOMICS_SHEET,
    WARNINGS_SHEET,
)
from .estimates import normalize_estimates
from .models import Activity, Assignment, BoardData, BoardSettings, Person, WarningMessage


BOARD_COLUMNS = ["key", "value"]
PEOPLE_COLUMNS = ["person_id", "name", "hours_per_day", "daily_cost", "active"]
ACTIVITY_COLUMNS = [
    "activity_id",
    "order",
    "name",
    "estimated_days",
    "estimated_hours",
    "max_hours_per_day",
    "daily_price",
    "status",
    "notes",
    "price_notes",
]
ASSIGNMENT_COLUMNS = ["activity_id", "person_id"]


def _blank_to_none(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    return value


def _as_str(value: Any, default: str = "") -> str:
    value = _blank_to_none(value)
    return default if value is None else str(value).strip()


def _as_float(value: Any, default: float | None = None) -> float | None:
    value = _blank_to_none(value)
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _as_int(value: Any, default: int | None = None) -> int | None:
    number = _as_float(value, None)
    if number is None:
        return default
    return int(number)


def _as_bool(value: Any, default: bool = True) -> bool:
    value = _blank_to_none(value)
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().upper() in {"TRUE", "YES", "Y", "1"}


def _as_date(value: Any, default: date | None = None) -> date:
    value = _blank_to_none(value)
    if value is None:
        return default or date.today()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return default or date.today()
    return parsed.date()


def _read_sheet(path: Path, sheet_name: str, columns: list[str]) -> tuple[pd.DataFrame, WarningMessage | None]:
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=object)
    except ValueError:
        return pd.DataFrame(columns=columns), WarningMessage("WARNING", "MISSING_SHEET", f"Missing sheet '{sheet_name}', using an empty default.")
    for column in columns:
        if column not in df.columns:
            df[column] = None
    return df[columns], None


def _settings_to_df(settings: BoardSettings) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"key": "board_name", "value": settings.board_name},
            {"key": "start_date", "value": settings.start_date.isoformat()},
            {"key": "hours_per_day", "value": settings.hours_per_day},
            {"key": "working_days", "value": ",".join(settings.working_days)},
        ],
        columns=BOARD_COLUMNS,
    )


def _people_to_df(people: list[Person]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "person_id": p.person_id,
                "name": p.name,
                "hours_per_day": p.hours_per_day,
                "daily_cost": p.daily_cost if p.daily_cost is not None else 0,
                "active": p.active,
            }
            for p in people
        ],
        columns=PEOPLE_COLUMNS,
    )


def _activities_to_df(activities: list[Activity]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "activity_id": a.activity_id,
                "order": a.order,
                "name": a.name,
                "estimated_days": a.estimated_days if a.estimated_days is not None else (a.estimated_hours / 8 if a.estimated_hours else None),
                "estimated_hours": a.estimated_hours,
                "max_hours_per_day": a.max_hours_per_day,
                "daily_price": a.daily_price if a.daily_price is not None else 0,
                "status": a.status,
                "notes": a.notes,
                "price_notes": a.price_notes,
            }
            for a in activities
        ],
        columns=ACTIVITY_COLUMNS,
    )


def _assignments_to_df(assignments: list[Assignment]) -> pd.DataFrame:
    return pd.DataFrame(
        [{"activity_id": a.activity_id, "person_id": a.person_id} for a in assignments],
        columns=ASSIGNMENT_COLUMNS,
    )


def _write_input_sheets(path: Path, board_data: BoardData) -> None:
    mode = "a" if path.exists() else "w"
    writer_kwargs: dict[str, object] = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"
    with pd.ExcelWriter(path, **writer_kwargs) as writer:
        _settings_to_df(board_data.settings).to_excel(writer, sheet_name=BOARD_SHEET, index=False)
        _people_to_df(board_data.people).to_excel(writer, sheet_name=PEOPLE_SHEET, index=False)
        _activities_to_df(board_data.activities).to_excel(writer, sheet_name=ACTIVITIES_SHEET, index=False)
        _assignments_to_df(board_data.assignments).to_excel(writer, sheet_name=ASSIGNMENTS_SHEET, index=False)


def create_new_board_file(path: Path, board_name: str, start_date: date) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    board = BoardData(
        settings=BoardSettings(board_name=board_name, start_date=start_date),
        people=[
            Person("P1", "Person 1", DEFAULT_HOURS_PER_DAY, True),
            Person("P2", "Person 2", DEFAULT_HOURS_PER_DAY, True),
        ],
        activities=[],
        assignments=[],
    )
    save_board(path, board)
    write_generated_sheets(
        path,
        pd.DataFrame(columns=["date", "person_id", "person_name", "activity_id", "activity_name", "hours"]),
        pd.DataFrame(columns=["metric", "value"]),
        pd.DataFrame(columns=["level", "code", "message"]),
        pd.DataFrame(
            columns=[
                "activity_id",
                "activity_name",
                "status",
                "estimated_hours",
                "estimated_person_days",
                "daily_price",
                "estimated_value",
                "allocated_hours",
                "allocated_person_days",
                "allocated_value",
                "remaining_allocated_hours_from_today",
                "remaining_allocated_value_from_today",
            ]
        ),
        pd.DataFrame(
            columns=[
                "person_id",
                "person_name",
                "daily_cost",
                "allocated_hours_total",
                "allocated_person_days_total",
                "estimated_delivery_cost",
                "delivered_hours_until_today",
                "delivered_cost_until_today",
                "remaining_hours_from_today",
                "remaining_delivery_cost_from_today",
                "allocated_until",
            ]
        ),
    )


def load_board(path: Path) -> BoardData:
    warnings: list[WarningMessage] = []
    board_df, warning = _read_sheet(path, BOARD_SHEET, BOARD_COLUMNS)
    if warning:
        warnings.append(warning)
    values = {str(row["key"]).strip(): row["value"] for _, row in board_df.iterrows() if _as_str(row.get("key"))}
    settings = BoardSettings(
        board_name=_as_str(values.get("board_name"), path.stem),
        start_date=_as_date(values.get("start_date"), date.today()),
        hours_per_day=_as_float(values.get("hours_per_day"), DEFAULT_HOURS_PER_DAY) or DEFAULT_HOURS_PER_DAY,
        working_days=tuple(
            part.strip().upper()
            for part in _as_str(values.get("working_days"), ",".join(DEFAULT_WORKING_DAYS)).split(",")
            if part.strip()
        )
        or DEFAULT_WORKING_DAYS,
    )

    people_df, warning = _read_sheet(path, PEOPLE_SHEET, PEOPLE_COLUMNS)
    if warning:
        warnings.append(warning)
    people = [
        Person(
            person_id=_as_str(row["person_id"]),
            name=_as_str(row["name"]),
            hours_per_day=_as_float(row["hours_per_day"], None),
            active=_as_bool(row["active"], True),
            daily_cost=_as_float(row["daily_cost"], 0),
        )
        for _, row in people_df.iterrows()
        if any(_blank_to_none(row[column]) is not None for column in PEOPLE_COLUMNS)
    ]

    activities_df, warning = _read_sheet(path, ACTIVITIES_SHEET, ACTIVITY_COLUMNS)
    if warning:
        warnings.append(warning)
    activities = []
    for _, row in activities_df.iterrows():
        if not any(_blank_to_none(row[column]) is not None for column in ACTIVITY_COLUMNS):
            continue
        activity_id = _as_str(row["activity_id"])
        estimated_days, estimated_hours, estimate_warnings = normalize_estimates(
            activity_id,
            row["estimated_days"],
            row["estimated_hours"],
        )
        warnings.extend(estimate_warnings)
        activities.append(
            Activity(
                activity_id=activity_id,
                order=_as_int(row["order"], None),
                name=_as_str(row["name"]),
                estimated_hours=estimated_hours,
                max_hours_per_day=_as_float(row["max_hours_per_day"], None),
                status=_as_str(row["status"], "PLANNED").upper(),
                notes=_as_str(row["notes"]),
                estimated_days=estimated_days,
                daily_price=_as_float(row["daily_price"], 0),
                price_notes=_as_str(row["price_notes"]),
            )
        )

    assignments_df, warning = _read_sheet(path, ASSIGNMENTS_SHEET, ASSIGNMENT_COLUMNS)
    if warning:
        warnings.append(warning)
    assignments = [
        Assignment(activity_id=_as_str(row["activity_id"]), person_id=_as_str(row["person_id"]))
        for _, row in assignments_df.iterrows()
        if any(_blank_to_none(row[column]) is not None for column in ASSIGNMENT_COLUMNS)
    ]

    return BoardData(settings=settings, people=people, activities=activities, assignments=assignments, warnings=warnings)


def save_board(path: Path, board_data: BoardData) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    _write_input_sheets(path, board_data)


def write_generated_sheets(
    path: Path,
    calendar_df: pd.DataFrame,
    dashboard_df: pd.DataFrame,
    warnings_df: pd.DataFrame,
    activity_economics_df: pd.DataFrame | None = None,
    person_economics_df: pd.DataFrame | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        create_new_board_file(path, path.stem, date.today())
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        calendar_df.to_excel(writer, sheet_name=CALENDAR_SHEET, index=False)
        dashboard_df.to_excel(writer, sheet_name=DASHBOARD_SHEET, index=False)
        warnings_df.to_excel(writer, sheet_name=WARNINGS_SHEET, index=False)
        if activity_economics_df is not None:
            activity_economics_df.to_excel(writer, sheet_name=ACTIVITY_ECONOMICS_SHEET, index=False)
        if person_economics_df is not None:
            person_economics_df.to_excel(writer, sheet_name=PERSON_ECONOMICS_SHEET, index=False)


def duplicate_board(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)


def list_board_files(boards_dir: Path) -> list[Path]:
    boards_dir.mkdir(parents=True, exist_ok=True)
    return sorted(boards_dir.glob("*.xlsx"))


def get_sheet_names(path: Path) -> list[str]:
    if not path.exists():
        return []
    return load_workbook(path, read_only=True).sheetnames
